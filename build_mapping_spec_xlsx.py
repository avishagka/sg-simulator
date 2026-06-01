from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("Mapping Specification - UniversalProviderDisplayPayload.xlsx")


HEADER_FILL = PatternFill("solid", fgColor="1F4D78")
SUBHEADER_FILL = PatternFill("solid", fgColor="E8EEF5")
LIGHT_FILL = PatternFill("solid", fgColor="F7FAFC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="DADCE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, freeze="A2"):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            else:
                cell.fill = WHITE_FILL if cell.row % 2 else LIGHT_FILL
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 70))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(max_len + 4, 55))


def add_sheet(wb, title, headers, rows, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, freeze)
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    overview_headers = ["נושא", "הגדרה"]
    overview_rows = [
        ["שם סכמה", "UniversalProviderDisplayPayload"],
        ["גרסה", "1.0"],
        ["מטרת המסמך", "מיפוי מה-JSON המקורי לשדות פיילואד ה-Agent"],
        ["תחולת v1", "רופאים בלבד (doctor)"],
        ["מבנה Response", "metadata, results[]"],
        ["Request", "מינימלי: requestId + query. endpoint קובע client metadata"],
        ["locale", "לא נשלח בבקשה; מוחזר ב-metadata.locale"],
        ["תוכן", "ב-v1 hardcoded ב-Agent; בעתיד Umbraco לפי contentKeys בכל רכיב. אין מילון תוכן מרכזי בפלט."],
        ["אחריות לוגיקה", "Agent/Orchestrator"],
        ["אחריות קליינט", "תוצאות חיפוש: רינדור לפי presentation.searchResult.componentIds. פרטי רופא: מיפוי רכיבי מידע לקומפוננטות UI ולאזורי תצוגה בצד הקליינט."],
    ]
    add_sheet(wb, "Overview", overview_headers, overview_rows)

    endpoint_headers = ["Endpoint", "clientType", "platform", "channel", "השפעה על אנשי קשר"]
    endpoint_rows = [
        ["אפליקציה", "mobile", "app", "memberApp", "טלפון/נייד/זימון תורים: tel; דוא\"ל: mailto; פקס/אחר: ללא פעולה"],
        ["אונליין", "web", "web", "online", "טלפון/נייד/זימון תורים: copy; דוא\"ל: mailto; פקס/אחר: ללא פעולה"],
        ["מל\"ה", "internalSystem", "web", "mlh", "טלפון/נייד/זימון תורים: copy; דוא\"ל: mailto; פקס/אחר: ללא פעולה"],
        ["מדריך השירותים", "web", "web", "serviceGuide", "טלפון/נייד/זימון תורים: copy; דוא\"ל: mailto; פקס/אחר: ללא פעולה"],
    ]
    add_sheet(wb, "Endpoint Mapping", endpoint_headers, endpoint_rows)

    mapping_headers = [
        "אזור",
        "Payload Path",
        "שם שדה",
        "סוג",
        "מקור JSON מקורי",
        "לוגיקה / חישוב",
        "מחושב?",
        "קבוע / החלטה עסקית?",
        "contentKeys.title / content ref",
        "הערות למפתח",
    ]

    rows = [
        ["metadata", "metadata.requestId", "requestId", "string", "Request או יצירת שרת", "מזהה בקשה; בסימולטור simulator-request", "כן", "לא", "", "במימוש אמיתי להעביר או לייצר בצד שרת"],
        ["metadata", "metadata.payloadVersion", "payloadVersion", "string", "קבוע", "1.0", "לא", "כן", "", ""],
        ["metadata", "metadata.schemaName", "schemaName", "string", "קבוע", "UniversalProviderDisplayPayload", "לא", "כן", "", ""],
        ["metadata", "metadata.schemaVersion", "schemaVersion", "string", "קבוע", "1.0", "לא", "כן", "", ""],
        ["metadata", "metadata.supportedEntityTypes", "supportedEntityTypes", "array<string>", "החלטה עסקית", "[doctor]", "לא", "כן", "", "v1 תומך רופאים בלבד"],
        ["metadata", "metadata.entityTypeVersions.doctor", "doctor", "string", "החלטה עסקית", "1.0", "לא", "כן", "", ""],
        ["metadata", "metadata.generatedAt", "generatedAt", "datetime", "Agent", "זמן יצירת הפלט בפורמט ISO", "כן", "לא", "", ""],
        ["metadata", "metadata.locale", "locale", "object", "החלטה עסקית", "{language: he, country: IL, direction: rtl}", "לא", "כן", "", "מוחזר בפלט בלבד"],
        ["metadata", "metadata.client", "client", "object", "endpoint/orchestrator", "נגזר מסוג endpoint", "כן", "לא", "", "הקליינט לא שולח clientContext"],
        ["metadata", "metadata.content", "content", "object", "החלטה עסקית", "mode=hardcoded, provider=agent, futureProvider=umbraco, contentVersion=hc-v1", "לא", "כן", "", "הכנה לאומברקו; לא לייצר מילון keys מרכזי"],
        ["metadata", "metadata.businessRules", "businessRules", "object", "Agent", "rulesetVersion=provider-search-v1, applied=true", "כן", "כן", "", ""],
        ["metadata", "metadata.sourceData", "sourceData", "object", "Agent", "sourceSystem=serviceGuide, sourcePayloadType=originalJson", "כן", "כן", "", ""],
        ["response", "results[]", "results", "array<object>", "תוצאות טרנספורמציה", "כל אובייקט מקור תקין הופך לרשומה אחת במערך results", "כן", "לא", "", "כל result מכיל entity, presentation, components בלבד. אין metadata בתוך result"],
        ["entity", "entity.isVisible", "isVisible", "boolean", "קבוע", "true עבור רשומת מקור תקינה", "לא", "כן", "", ""],
        ["entity", "entity.entityType", "entityType", "string", "chapter_code", "001/01 => doctor; 003/03 => therapist; אחרת other", "כן", "לא", "", "ב-v1 מצופה doctor"],
        ["entity", "entity.id", "id", "string", "entity_id, service_provider_id, object_id", "נלקח לפי זמינות בסדר זה", "כן", "לא", "", ""],
        ["entity", "entity.sourceType", "sourceType", "string", "chapter_code", "קוד סוג ישות מהמקור", "לא", "לא", "", ""],
        ["entity", "entity.sourceCode", "sourceCode", "string", "object_id", "קוד מקור/אובייקט", "לא", "לא", "", ""],
        ["entity", "entity.name", "name", "string", "service_name", "שם השירות", "לא", "לא", "", "חשוב: name צריך להיות service_name"],
        ["entity", "entity.displayName", "displayName", "string", "service_provider_name, service_name", "service_provider_name; fallback service_name", "כן", "לא", "", "שם לתצוגה"],
        ["entity", "entity.subtitle", "subtitle", "string", "treat_area_string", "ערך מקור אם קיים", "לא", "לא", "", ""],
        ["entity", "entity.profileUrl", "profileUrl", "string/null", "url", "ערך מקור אם קיים", "לא", "לא", "", ""],
        ["entity.image", "entity.image.type", "type", "string", "קבוע", "avatar", "לא", "כן", "", ""],
        ["entity.image", "entity.image.gender", "gender", "string/null", "service_provider_gender", "ערך מקור", "לא", "לא", "", ""],
        ["entity.image", "entity.image.avatarType", "avatarType", "string", "service_provider_gender", "אישה/F/female => female; גבר/M/male => male; אחרת default", "כן", "לא", "", ""],
        ["entity.image", "entity.image.url", "url", "string/null", "החלטה עסקית", "null ב-v1", "לא", "כן", "", ""],
        ["entity.image", "entity.image.alt", "alt", "string", "service_provider_name, service_name", "service_provider_name; fallback service_name", "כן", "לא", "", ""],
        ["specializations", "components.specializations.values[].value", "value", "string", "specialization_list", "כל ערך לא ריק הופך לערך בפלט; fallback ל-service_name_doc אם הרשימה חסרה", "כן", "לא", "provider.specializations.title", "מומחיות ← specialization_list"],
        ["professionalizations", "components.professionalizations.values[].value", "value", "string", "treat_area_list[].treat_area", "כל treat_area לא ריק הופך לערך בפלט", "כן", "לא", "provider.professionalizations.title", "התמקצעות ← treat_area_list"],
        ["languages", "components.languages.values[].value", "value", "string", "languages[].language", "מיון לפי seqnr_languages; fallback לפיצול language_list", "כן", "לא", "provider.languages.title", ""],
        ["address", "components.address.displayAddress", "displayAddress", "string/null", "full_address או city_name/street_name/house_number", "full_address קודם; אחרת חיבור חלקי כתובת קיימים", "כן", "לא", "provider.address.title", ""],
        ["address", "components.address.accessibility", "accessibility", "object", "accessibility_bool, accessibility", "אם accessibility_bool=true: טקסט accessibility או fallback נגיש", "כן", "לא", "", ""],
        ["remarks", "components.remarks.items[]", "items", "array<object>", "remarks[]", "רק remark_text לא ריק; מיפוי לפי remark_message_type_code", "כן", "לא", "", "אין כותרת לרכיב remarks"],
        ["remarks", "components.remarks.items[].sortOrder", "sortOrder", "number", "remark_message_type_code, remark_seqnr, remark_line_number", "sortBase*1000 + remark_seqnr*10 + remark_line_number", "כן", "לא", "", "2011 הוא מיון מורכב ולא מספר רץ פשוט"],
        ["icons", "components.icons.items[type=accessibility]", "accessibility icon", "object", "accessibility_bool, accessibility", "אם accessibility_bool=true מוסיפים אייקון נגישות", "כן", "לא", "", "אין כותרת לרכיב icons"],
        ["icons", "components.icons.items[type=limitedMembership]", "limitedMembership", "object", "noMember/no_member/no_member_icon/limited_memb_icon", "אם אחד מהשדות true מוסיפים לא מקבל חברים חדשים", "כן", "לא", "", ""],
        ["icons", "components.icons.items[type=shabanConsultation]", "shabanConsultation", "object", "shaban_type", "1 או 3 => ייעוץ בתשלום נוסף", "כן", "לא", "", "גם אם shaban_ind=false"],
        ["icons", "components.icons.items[type=shabanTreatments]", "shabanTreatments", "object", "shaban_type", "2 או 3 => ניתוחים וטיפולים בתשלום נוסף", "כן", "לא", "", "גם אם shaban_ind=false"],
        ["facil", "components.facil", "facil", "object", "החלטה עסקית", "רכיב עתידי; isVisible=false; items=[]", "לא", "כן", "provider.facil.title", ""],
        ["absences", "components.absences.items[]", "items", "array<object>", "absence[]", "המרת תאריכים, בניית displayText, מיון לפי תאריך התחלה", "כן", "לא", "provider.absences.title", ""],
        ["absences", "components.absences.items[].replacementEntity", "replacementEntity", "object", "subtitute[] או substitute[]", "התאמה לפי אותם תאריכי התחלה וסיום", "כן", "לא", "", ""],
        ["schedule", "components.schedule.groups[]", "groups", "array<object>", "schedule[]", "קיבוץ לפי schedule_type ומיון עסקי", "כן", "לא", "provider.schedule.title", ""],
        ["schedule", "components.schedule.groups[].displayDays", "displayDays", "array<object>", "days[] מחושב", "איחוד ימים רצופים עם אותן שעות ואותן הערות בדיוק", "כן", "לא", "", "הקליינט משתמש ב-displayDays להצגה"],
        ["schedule", "components.schedule.groups[].display.defaultOpen", "defaultOpen", "boolean", "schedule_type", "type=reception פתוח; אחרים סגורים", "כן", "לא", "", ""],
        ["services", "components.services.groups[]", "groups", "array<object>", "serv_treats[]", "קיבוץ לפי cpt_pubt או 00000; מיון לפי הופעה ראשונה", "כן", "לא", "provider.services.title", "ניתוחים וטיפולים"],
        ["services", "components.services.groups[].items[].name", "name", "string", "sg_treat_name/cell_treat_name/hebrew_text/hebrew_service_name", "שם טיפול לפי סדר עדיפות", "כן", "לא", "", ""],
        ["services", "components.services.groups[].items[].coverageType", "coverageType", "string", "treat_is_personal", "true => supplementary; אחרת basket", "כן", "לא", "", ""],
        ["services", "components.services.groups[].items[].coverageText", "coverageText", "string", "treat_is_personal", "true => משלים; אחרת בסל", "כן", "לא", "provider.services.coverage.basket / provider.services.coverage.supplementary", ""],
        ["services", "components.services.groups[].display", "display", "object", "items.length", "עד 2 פריטים פתוח ללא overflow; 3+ סגור עם previewRows=2", "כן", "לא", "", ""],
        ["appointments", "components.appointments", "appointments", "object", "badge limitedMembership", "רכיב מסך זימון תורים; מציג הודעות בלבד ב-v1", "כן", "לא", "provider.appointments.title", "לא לערבב עם contacts type=appointments"],
        ["contacts", "components.contacts.items[]", "items", "array<object>", "contact[]", "קיבוץ לפי contact_code, מניעת כפילויות, מיון עסקי", "כן", "לא", "provider.contacts.title", "אמצעי תקשורת רכיב מסך"],
        ["contacts", "components.contacts.items[].values[]", "values", "array<object>", "contact_detail", "כל מספר/מייל/פקס נשמר כערך נפרד; לא מחברים מספרים", "כן", "לא", "", ""],
        ["contacts", "components.contacts.items[].values[].preferredAction", "preferredAction", "object/null", "endpoint + contact type", "אפליקציה tel; אונליין/מל\"ה/מדריך copy לטלפונים; email mailto; fax/other null", "כן", "לא", "", ""],
        ["contacts", "components.contacts.messages[]", "messages", "array<object>", "contact type=email", "אם יש דוא\"ל נוצרת הודעת disclaimer", "כן", "לא", "provider.contacts.emailDisclaimer", ""],
        ["resume", "components.resume.sections[type=license]", "license section", "object", "license_number", "אם קיים נוצר סקשן רישיון", "כן", "לא", "provider.resume.license.title", "resume מחליף details"],
        ["resume", "components.resume.sections[type=resume]", "resume sections", "array<object>", "resume[]", "רק זוגות label/value תקינים; fallback title פרטי השכלה ומומחיות", "כן", "לא", "provider.resume.sectionTitle", ""],
        ["presentation", "presentation.searchResult.componentIds", "componentIds", "array<string>", "החלטה עסקית", "entity, specializations, professionalizations, languages, remarks, icons, address, absences", "לא", "כן", "provider.searchResult.title", "סדר תצוגה מחייב לכרטיס תוצאת חיפוש"],
        ["presentation", "presentation.details.layoutPolicy", "layoutPolicy", "string", "החלטה עסקית", "clientManaged", "לא", "כן", "provider.details.screenTitle", "במסך פרטי רופא הקליינט אחראי למיפוי רכיבי המידע לקומפוננטות UI ולאזורי התצוגה"],
        ["presentation", "presentation.details.availableComponentIds", "availableComponentIds", "array<string>", "החלטה עסקית", "remarks, entity, specializations, professionalizations, languages, icons, absences, schedule, address, facil, services, appointments, contacts, resume", "לא", "כן", "provider.details.screenTitle", "רשימת רכיבי מידע זמינים בלבד; אין משמעות של סדר, מיקום או שיוך לאזור layout"],
    ]
    def prefix_result_path(row):
        updated = list(row)
        path = str(updated[1])
        if updated[0] not in ("metadata", "response") and not path.startswith("results[]"):
            updated[1] = f"results[].{path}"
        return updated

    component_contract_rows = [
        ["components", "results[].components.<component>.id", "id", "string", "קבוע לפי שם הרכיב", "חובה בכל component לפי OpenAPI והסימולטור", "לא", "כן", "", "לדוגמה: schedule.id=schedule"],
        ["components", "results[].components.<component>.isVisible", "isVisible", "boolean", "מחושב מתוכן הרכיב", "חובה בכל component; false אם אין תוכן להצגה או רכיב עתידי", "כן", "לא", "", ""],
        ["components", "results[].components.<component>.contentKeys.title", "contentKeys.title", "string", "מפתח תוכן", "חובה ברכיבים עם title; מבנה object לפי OpenAPI", "לא", "כן", "", "להשתמש ב-contentKeys בלבד"],
        ["textList", "results[].components.specializations.values[].sortOrder", "sortOrder", "number", "סדר ערכים לאחר מיפוי", "סדר פנימי בתוך values[]", "כן", "לא", "", ""],
        ["textList", "results[].components.professionalizations.values[].sortOrder", "sortOrder", "number", "סדר ערכים לאחר מיפוי", "סדר פנימי בתוך values[]", "כן", "לא", "", ""],
        ["textList", "results[].components.languages.values[].sortOrder", "sortOrder", "number", "seqnr_languages או סדר fallback", "סדר פנימי בתוך values[]", "כן", "לא", "", ""],
        ["contacts", "results[].components.contacts.items[].sortOrder", "sortOrder", "number", "contact_code", "סדר סוג קשר פשוט: phone=1, appointments=2, fax=3, mobile=4, email=5, other=999", "כן", "לא", "", ""],
        ["remarks", "results[].components.remarks.items[].sourceTypeCode", "sourceTypeCode", "string", "remark_message_type_code", "שם השדה לפי הסימולטור וה-OpenAPI", "לא", "לא", "", "לא sourceCode"],
        ["contacts", "results[].components.contacts.items[].values[].sortOrder", "sortOrder", "number", "סדר ערכים בתוך סוג קשר", "כל מספר/מייל/פקס מקבל sortOrder נפרד", "כן", "לא", "", ""],
        ["contacts", "results[].components.contacts.items[].values[].preferredAction.type", "type", "string", "endpoint + contact type", "tel/copy/mailto או null ברמת preferredAction", "כן", "לא", "", ""],
        ["contacts", "results[].components.contacts.items[].values[].preferredAction.uri", "uri", "string", "value", "עבור tel/mailto בלבד: tel:<value> או mailto:<value>", "כן", "לא", "", "לא להשתמש ב-value עבור tel/mailto"],
        ["contacts", "results[].components.contacts.items[].values[].preferredAction.value", "value", "string", "value", "עבור copy בלבד", "כן", "לא", "", ""],
        ["schedule", "results[].components.schedule.groups[].days[].dayCode", "dayCode", "number", "availability_week_day", "יום 1-7 לפי הסימולטור", "כן", "לא", "", "לא dayOfWeek"],
        ["schedule", "results[].components.schedule.groups[].days[].dayLabel", "dayLabel", "string", "dayCode", "א׳, ב׳, ג׳, ד׳, ה׳, ו׳, שבת", "כן", "לא", "", "לא dayName"],
        ["schedule", "results[].components.schedule.groups[].days[].isVisible", "isVisible", "boolean", "rows.length", "true אם יש rows ליום", "כן", "לא", "", ""],
        ["schedule", "results[].components.schedule.groups[].days[].rows[]", "rows", "array<object>", "schedule[]", "כל טווח שעות נשמר בתוך rows[] של אותו יום", "כן", "לא", "", "לא לשטח startTime/endTime ישירות על day"],
        ["schedule", "results[].components.schedule.groups[].displayDays[].dayCodes", "dayCodes", "array<number>", "days[] מחושב", "ימי המקור שאוחדו לתצוגה", "כן", "לא", "", ""],
        ["schedule", "results[].components.schedule.groups[].displayDays[].dayLabel", "dayLabel", "string", "days[] מחושב", "תווית יום יחיד או טווח, לדוגמה א׳-ב׳", "כן", "לא", "", "לא label"],
        ["schedule", "results[].components.schedule.groups[].displayDays[].isRange", "isRange", "boolean", "days[] מחושב", "true אם אוחדו יותר מיום אחד", "כן", "לא", "", ""],
        ["schedule", "results[].components.schedule.groups[].displayDays[].rows[]", "rows", "array<object>", "rows[] של היום הראשון בקבוצה", "אותו מבנה rows כמו days[].rows[]", "כן", "לא", "", "לא לשטח startTime/endTime ישירות על displayDay"],
        ["schedule", "results[].components.schedule.groups[].display.previewRows", "previewRows", "number/null", "החלטת תצוגה", "קיים לפי OpenAPI גם אם null", "כן", "לא", "", ""],
        ["schedule", "results[].components.schedule.groups[].display.hasOverflow", "hasOverflow", "boolean", "החלטת תצוגה", "קיים לפי OpenAPI", "כן", "לא", "", ""],
        ["resume", "results[].components.resume.sections[].sortOrder", "sortOrder", "number", "license/resume_code/index", "סדר פנימי של sections", "כן", "לא", "", ""],
        ["resume", "results[].components.resume.sections[].items[].sortOrder", "sortOrder", "number", "סדר השדות במקור", "סדר פנימי של items", "כן", "לא", "", ""],
    ]

    rows = [prefix_result_path(row) for row in rows] + component_contract_rows
    add_sheet(wb, "Field Mapping", mapping_headers, rows)

    contacts_headers = ["contact_code", "type", "label", "sortOrder", "אפליקציה", "אונליין", "מל\"ה", "מדריך השירותים"]
    contacts_rows = [
        ["01", "phone", "טלפון", 1, "tel", "copy", "copy", "copy"],
        ["02", "appointments", "זימון תורים", 2, "tel", "copy", "copy", "copy"],
        ["03", "fax", "פקס", 3, "none", "none", "none", "none"],
        ["04", "mobile", "נייד", 4, "tel", "copy", "copy", "copy"],
        ["05", "email", "דוא\"ל", 5, "mailto", "mailto", "mailto", "mailto"],
        ["other", "other", "contact_desc או אחר", 999, "none", "none", "none", "none"],
    ]
    add_sheet(wb, "Contact Actions", contacts_headers, contacts_rows)

    schedule_headers = ["schedule_type", "type", "title", "sortOrder", "defaultOpen", "לוגיקה"]
    schedule_rows = [
        [1, "reception", "שעות פעילות הרופא", 1, True, "פתוח כברירת מחדל"],
        [2, "appointments", "זימון תורים", 2, False, "סגור כברירת מחדל"],
        [5, "additional", "נוסף", 3, False, "סגור כברירת מחדל"],
        [14, "phone", "שעות פעילות טלפונית", 4, False, "סגור כברירת מחדל"],
        [15, "teamReception", "קבלת קהל צוות", 5, False, "סגור כברירת מחדל"],
        ["other", "other", "schedule_desc או אחר", 999, False, "fallback"],
    ]
    add_sheet(wb, "Schedule Rules", schedule_headers, schedule_rows)

    remarks_headers = ["remark_message_type_code", "type", "targetArea", "displayStyle", "sortBase", "isVisible", "severity"]
    remarks_rows = [
        ["0010 / 00010 / 10", "emergency", "pageTop", "highlight", 1, True, "green"],
        ["0004", "name", "entity", "inline", 2, False, ""],
        ["0006", "contact", "contacts", "inline", 3, False, ""],
        ["0002", "address", "locations", "inline", 4, False, ""],
        ["0003", "schedule", "availability", "inline", 5, False, ""],
        ["0005", "occupation", "entity", "regular", 6, False, ""],
        ["0008", "clinicServices", "services", "regular", 7, False, ""],
        ["0020 / 0001", "additionalInfo", "details", "regular", 8, False, ""],
        ["other", "other", "details", "regular", 999, False, ""],
    ]
    add_sheet(wb, "Remark Rules", remarks_headers, remarks_rows)

    validation_headers = ["בדיקה", "ציפייה"]
    validation_rows = [
        ["chapter_code=001", "entityType=doctor"],
        ["specialization_list", "מאכלס components.specializations"],
        ["treat_area_list", "מאכלס components.professionalizations"],
        ["languages", "ממוינות לפי seqnr_languages"],
        ["shaban_type=2", "יוצר תגית ניתוחים וטיפולים גם אם shaban_ind=false"],
        ["shaban_type=3", "יוצר שתי תגיות שב\"ן"],
        ["remarks 0004", "ממופה ל-type=name"],
        ["שני מספרי זימון תורים", "נשמרים כשני values נפרדים; לא מחברים למחרוזת אחת"],
        ["אפליקציה + appointments contact", "preferredAction.type=tel לכל מספר"],
        ["אונליין/מל\"ה/מדריך + appointments contact", "preferredAction.type=copy לכל מספר"],
        ["contacts legacy actions", "לא קיימים action/desktopAction/mobileAction; משתמשים רק ב-preferredAction"],
        ["דוא\"ל", "mailto + disclaimer"],
        ["פקס/אחר", "preferredAction=null"],
        ["schedule rows", "ממוינים לפי שעת התחלה"],
        ["ימים רצופים זהים", "מאוחדים ב-displayDays"],
        ["ימים לא רצופים", "לא מאוחדים גם אם שעות זהות"],
        ["services עם 2 פריטים", "defaultOpen=true, hasOverflow=false"],
        ["services עם 3 פריטים", "defaultOpen=false, hasOverflow=true, previewRows=2"],
    ]
    add_sheet(wb, "Validation Checklist", validation_headers, validation_rows)

    wb.save(OUTPUT)

    # Structural verification.
    check = load_workbook(OUTPUT)
    expected = {
        "Overview",
        "Endpoint Mapping",
        "Field Mapping",
        "Contact Actions",
        "Schedule Rules",
        "Remark Rules",
        "Validation Checklist",
    }
    missing = expected - set(check.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    if check["Field Mapping"].max_row < 50:
        raise RuntimeError("Field Mapping sheet is unexpectedly short")
    print(OUTPUT)


if __name__ == "__main__":
    main()
